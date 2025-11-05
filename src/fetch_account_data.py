import datetime
import logging
import os

import openpyxl
from plombery import task, get_logger

from src.utils import logger
from src.rules.CompanyEnum import CompanyNameEnum
from src.sms.sms_client import SMSClient
from src.utils import ConfigLoader, dbutils

config = ConfigLoader.ConfigLoader()
sms_client = SMSClient(base_url=config.get_sms_api())
job_name = "fetch_account_data"
job_config = config.get_config_by_job(job_name)
logger = logging.getLogger(__name__)


def fill_sql(job_config):
    """填充config中的sql参数并返回最终可执行的sql"""
    sql_template = job_config.get("sql")

    # 获取当前日期
    today = datetime.date.today()
    current_month = today.strftime("%Y%m")  # 202509 -> mr_month 格式

    # 上一年同月
    last_year_date = today.replace(year=today.year - 1)
    last_year_month = last_year_date.strftime("%Y-%m")

    # 上一个月
    first_day_of_this_month = today.replace(day=1)
    previous_month_date = first_day_of_this_month - datetime.timedelta(days=1)
    previous_month = previous_month_date.strftime("%Y-%m")

    # 填充 SQL 模板
    sql_filled = sql_template.format(
        previous=previous_month_date.year,
        previous_month=previous_month,
        last_year=last_year_date.year,
        last_year_month=last_year_month,
        last_year_date =last_year_date,
        previous_year_date =previous_month_date,
        work_day=today.day # TODO:后续取计划管理的工作日，目前暂时按自然日取
    )
    return sql_filled

def fetch_data_by_company(company: str):
    """按分公司获取数据"""
    company_config =  config.get_database(company)
    db = dbutils.DBUtils(company_config)
    db.connect()

    # 获取公司对应数据库配置（如果有不同公司）
    sql = fill_sql(job_config)
    users = db.query(sql)
    db.close()
    # 返回格式 [{'count': 139930, 'seq': 1}, {'count': 0, 'seq': 2}, {'count': 134194, 'seq': 3}]
    # 1-本周期 2-上周期 3-去年同期
    data = {
        '本周期户表出账(支)': users[0]['count'] if len(users) > 0 else -1,
        '本周期户表应出账(支)': users[1]['count'] if len(users) > 1 else -1,
        '本周期大路表出账(支)': users[2]['count'] if len(users) > 0 else -1,
        '本周期大路表应出账(支)': users[3]['count'] if len(users) > 1 else -1,
        '上周期户表出账(支)': users[4]['count'] if len(users) > 0 else -1,
        '去年周期户表出账(支)': users[5]['count'] if len(users) > 1 else -1,
        '上周期大路表出账(支)': users[6]['count'] if len(users) > 1 else -1,
        '去年周期大路表出账(支)': users[7]['count'] if len(users) > 1 else -1,
    }

    res = {
        'company': company,
        'data': data
    }

    logger.info("处理后数据：%s", res)
    return res


def build_excel(all_company_data,template_name="远传表出账明细.xlsx", output_path="远传表出账明细.xlsx"):
    '''查询出来的结果写入到excel'''

    # 1. 定义数据键名与Excel行号的映射关系（基于模板图片）
    # 假设模板表格的“类目”在 A1，数据从 A2 开始。
    # 模板行名:           数据字典键名:                     行号 (row index)
    DATA_ROW_MAP = {
        '本周期户表应出账': '本周期户表应出账(支)',  # 第 2 行
        '本周期户表实际出账': '本周期户表出账(支)',  # 第 3 行
        '上周期户表实际出账': '上周期户表出账(支)',  # 第 4 行
        '去年同期户表实际出账': '去年周期户表出账(支)',  # 第 5 行
        '本周期大路表应出账': '本周期大路表应出账(支)',  # 第 6 行
        '本周期大路表实际出账': '本周期大路表出账(支)',  # 第 7 行
        '上周期大路表实际出账': '上周期大路表出账(支)',  # 第 8 行
        '去年同期大路表实际出账': '去年周期大路表出账(支)',  # 第 9 行
    }

    # 将字典值（数据键名）作为映射的查找键
    DATA_KEY_TO_ROW = {v: k for k, v in DATA_ROW_MAP.items()}

    # Excel 的起始列 (B列，列索引为 2)
    START_COL_INDEX = 2

    try:
        # 2. 加载模板
        current_dir = os.getcwd()
        # 构建模板文件的完整路径：run/template/远传出账明细模板.xlsx
        template_path = os.path.join(current_dir, 'template', template_name)
        if not os.path.exists(template_path):
            logger.error(f"模板文件未找到: {template_path}。请确保模板存在。")
            # 如果模板不存在，可以创建基础结构 (简化处理)
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.cell(row=1, column=1, value="远传出账明细")
            ws.cell(row=2, column=1, value="类目")
            # 写入类目名称
            for i, category in enumerate(DATA_ROW_MAP.keys()):
                ws.cell(row=i + 3, column=1, value=category)
            # 为了匹配图片，重新调整 row index
            for i, category in enumerate(DATA_ROW_MAP.keys()):
                ws.cell(row=i + 2, column=1, value=category)
        else:
            wb = openpyxl.load_workbook(template_path)

        ws = wb.active
        for i, data_obj in enumerate(all_company_data):

            company_code = data_obj.get('company')
            company_name = CompanyNameEnum.get_name(company_code)  # 写入列
            data_to_write = data_obj.get('data', {})

            # 当前写入的列索引：从 B 列 (2) 开始，随着 i 递增
            current_col = START_COL_INDEX + i

            # 5. 写入公司名称 (列标题)
            # 写入公司名称到第 2 行 (Row 2) 作为列头
            ws.cell(row=2, column=current_col, value=company_name)

            # 6. 遍历映射，写入该公司的各项数据
            for key_in_data, row_name in DATA_KEY_TO_ROW.items():
                if key_in_data in data_to_write:
                    try:
                        list_of_row_names = list(DATA_ROW_MAP.keys())
                        index = list_of_row_names.index(row_name)
                        # 实际行号是索引 + 3 (A1 标题, A2 类目/公司名, A3 是第一项数据)
                        row_index = index + 3

                    except ValueError:
                        continue

                    # 写入数据到对应的行和当前的列 (current_col)
                    value = data_to_write[key_in_data]
                    ws.cell(row=row_index, column=current_col, value=value)

            logger.info(f"已将 {company_name} 数据写入 Excel 第 {current_col} 列 ({chr(64 + current_col)} 列)。")

            # 7. 保存 Excel 文件 (在循环结束后保存)
        wb.save(output_path)
        logger.info(f"Excel 文件已成功生成并保存到: {output_path}")
        return True

    except Exception as e:
        logger.error(f"生成 Excel 出现错误: {e}")
        return False



def build_sms_message(all_data):
    message = "【生产环境】截至目前本月远传出账情况："
    for data in all_data:
        message += '\n'
        message += CompanyNameEnum.get_name(data['company'].upper()) + ':' + str(data['data']).replace('{', '').replace('}',
                                                                                                             '').replace(
            '\'', '')
    logger.debug(message)
    return message

# 注册到plombery
@task
def fetch_account_data_job():
    """定时任务：发送生产环境远传出账生成情况"""
    logger = get_logger()
    companies = job_config.get("companies")
    res = []
    # 获取所有分公司数据
    for company in companies:
        data = fetch_data_by_company(company)
        res.append(data)
    if res:
        message = build_sms_message(res)
        sms_client.send_sms(phones=job_config['phones'],content=message,logger=logger)
        build_excel(all_company_data=res)
    else:
        logger.warning("未获取到任何公司数据，Excel生成跳过。")
        logger.warning("未获取到任何公司数据，短信发送跳过。")


if __name__ == "__main__":
    # 手动触发测试
    fetch_account_data_job()
    # build_sms_message()
    # fill_sql(job_config)